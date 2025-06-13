import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import cohere
import numpy as np
from openpyxl import load_workbook
import os
import threading
from datetime import datetime
import re
from rapidfuzz import fuzz

# --- CONFIGURABLE CONSTANTS ---
EMBEDDING_MODEL = "embed-v4.0"
EMBEDDING_BATCH_SIZE = 96        # Cohere supports up to 96 texts per embed call
EMBEDDING_DIMENSION = 1536       # Choose from 256, 512, 1024, or 1536
FUZZY_THRESHOLD = 0.4            # Embedding score threshold to trigger fuzzy-text fallback
FALLBACK_CANDIDATES = 5          # Number of top embedding candidates for fuzzy re-rank

class PricelistMatcherApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pricelist Matching Application (Cohere Embed v4.0)")
        self.root.geometry("670x550")

        # -- default values; overrideable in UI --
        self.pricelist_path = tk.StringVar(value=r"C:\\Users\\aymanhofi\\OneDrive\\Side Projects\\MJD\\Files\\Cleaned files\\MJD-Pricelist_Cleaned.xlsx")
        self.inquiry_path   = tk.StringVar(value=r"C:\\Users\\aymanhofi\\OneDrive\\Side Projects\\MJD\\Files\\Cleaned files\\ETON Inqury Emty pages.xlsx")
        self.api_key_var    = tk.StringVar(value="NMRRRXIjenLXk4nQhNc3PYhZ13vdMlVLYGRYqkto")
        self.output_folder  = tk.StringVar(value=r"C:\\Users\\aymanhofi\\OneDrive\\Side Projects\\MJD\\Files\\Cleaned files")
        self.use_fuzzy      = tk.BooleanVar(value=False)
        self.use_taxonomy   = tk.BooleanVar(value=False)
        self.client         = None

        self.build_widgets()

    def build_widgets(self):
        frm = tk.Frame(self.root, padx=10, pady=10)
        frm.pack(fill=tk.BOTH, expand=True)

        # Pricelist chooser
        tk.Label(frm, text="Pricelist Excel:").grid(row=0, column=0, sticky="e")
        tk.Entry(frm, textvariable=self.pricelist_path, width=45).grid(row=0, column=1, padx=5)
        tk.Button(frm, text="Browse...", command=lambda: self.pricelist_path.set(
            filedialog.askopenfilename(title="Select Pricelist File", filetypes=[("Excel files","*.xlsx *.xls")])
        )).grid(row=0, column=2, padx=5)

        # Inquiry chooser
        tk.Label(frm, text="Inquiry Excel:").grid(row=1, column=0, sticky="e")
        tk.Entry(frm, textvariable=self.inquiry_path, width=45).grid(row=1, column=1, padx=5)
        tk.Button(frm, text="Browse...", command=lambda: self.inquiry_path.set(
            filedialog.askopenfilename(title="Select Inquiry File", filetypes=[("Excel files","*.xlsx *.xls")])
        )).grid(row=1, column=2, padx=5)

        # API key
        tk.Label(frm, text="Cohere API Key:").grid(row=2, column=0, sticky="e")
        tk.Entry(frm, textvariable=self.api_key_var, show="*", width=45).grid(row=2, column=1, columnspan=2, sticky="w")

        # Output folder
        tk.Label(frm, text="Output Folder:").grid(row=3, column=0, sticky="e")
        tk.Entry(frm, textvariable=self.output_folder, width=45).grid(row=3, column=1, padx=5)
        tk.Button(frm, text="Browse...", command=lambda: self.output_folder.set(
            filedialog.askdirectory(title="Select Output Folder")
        )).grid(row=3, column=2, padx=5)

        # Feature toggles
        tk.Checkbutton(frm, text="Enable fuzzy-text fallback", variable=self.use_fuzzy).grid(row=4, column=1, sticky="w")
        tk.Checkbutton(frm, text="Copy Category/SubCategory", variable=self.use_taxonomy).grid(row=4, column=2, sticky="w")

        # Progress log
        tk.Label(frm, text="Progress:").grid(row=5, column=0, sticky="nw", pady=(10,0))
        self.log_box = scrolledtext.ScrolledText(frm, width=85, height=14, state="disabled", wrap=tk.WORD)
        self.log_box.grid(row=5, column=1, columnspan=2, pady=(10,0))

        # Process button
        self.process_btn = tk.Button(frm, text="Process", command=self.on_process_thread,
                                     bg="#4CAF50", fg="white", width=14)
        self.process_btn.grid(row=6, column=1, pady=12, sticky="w")

    def log(self, msg):
        self.log_box.config(state="normal")
        self.log_box.insert(tk.END, msg + "\n")
        self.log_box.see(tk.END)
        self.log_box.config(state="disabled")
        self.root.update_idletasks()

    def get_auto_output_path(self):
        folder = self.output_folder.get()
        if not folder:
            raise RuntimeError("Please specify an output folder.")
        now = datetime.now()
        filename = now.strftime("Output_%I-%M-%p_%m-%d-%y.xlsx")
        return os.path.join(folder, filename)

    def on_process_thread(self):
        t = threading.Thread(target=self.on_process)
        t.daemon = True
        t.start()

    def on_process(self):
        self.process_btn.config(state=tk.DISABLED)
        try:
            if not (self.pricelist_path.get() and self.inquiry_path.get()):
                raise RuntimeError("Please select both Pricelist and Inquiry files.")
            if not self.api_key_var.get().strip():
                raise RuntimeError("Please enter your Cohere API key.")
            if not self.output_folder.get():
                raise RuntimeError("Please specify an output folder.")

            self.client = cohere.ClientV2(api_key=self.api_key_var.get().strip())
            self.log("Initialized Cohere client.")

            output_path = self.get_auto_output_path()
            if os.path.exists(output_path):
                if not messagebox.askyesno("Overwrite?", f"Output file '{output_path}' exists. Overwrite?"):
                    self.log("Process cancelled by user.")
                    return

            self.log("Starting processing...")
            texts, descs, rates, cats, subs = load_pricelist_data(
                self.pricelist_path.get(), self.log, self.use_taxonomy.get()
            )
            wb_inq, items_to_fill, header_rows = load_inquiry_data(
                self.inquiry_path.get(), self.log
            )

            fill_inquiry_rates(
                client=self.client,
                wb_inq=wb_inq,
                items_to_fill=items_to_fill,
                pricelist_texts=texts,
                pricelist_simple_descs=descs,
                pricelist_rates=rates,
                pricelist_cats=cats,
                pricelist_subs=subs,
                header_rows=header_rows,
                model=EMBEDDING_MODEL,
                logger_fn=self.log,
                use_fuzzy=self.use_fuzzy.get(),
                use_taxonomy=self.use_taxonomy.get()
            )

            self.log(f"Saving output file as: {output_path}")
            wb_inq.save(output_path)
            wb_inq.close()
            self.log("Output file saved.")
            messagebox.showinfo("Success", f"Pricing completed.\nOutput saved to:\n{output_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log(f"Error: {e}")
        finally:
            self.process_btn.config(state=tk.NORMAL)

# --- CORE PROCESSING FUNCTIONS ---

def preprocess_text(s):
    if not s:
        return ""
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("mm.", "mm").replace("cm.", "cm")
    s = s.replace("r.c.c.", "rcc").replace("reinforced cement concrete", "rcc")
    return s

def load_pricelist_data(pricelist_path, logger_fn, include_taxonomy):
    logger_fn("Reading pricelist file...")
    wb_price = load_workbook(pricelist_path, read_only=True, data_only=True)
    texts, descs, rates, cats, subs = [], [], [], [], []
    for sheet in wb_price.worksheets:
        logger_fn(f"Processing pricelist sheet '{sheet.title}'...")
        num_cols = sheet.max_column
        desc_idx = rate_idx = cat_idx = sub_idx = None
        header_row_idx = None
        for r in range(1, 6):
            row_vals = next(sheet.iter_rows(min_row=r, max_row=r, max_col=num_cols, values_only=True))
            cleaned = [str(v).strip().lower() if isinstance(v, str) else "" for v in row_vals]
            if "description" in cleaned and "rate" in cleaned:
                desc_idx = cleaned.index("description")
                rate_idx = cleaned.index("rate")
                if include_taxonomy:
                    cat_idx = cleaned.index("category") if "category" in cleaned else None
                    sub_idx = cleaned.index("subcategory") if "subcategory" in cleaned else None
                header_row_idx = r
                headers = [str(v).strip() if v is not None else f"col{ci}" for ci, v in enumerate(row_vals)]
                break
        if header_row_idx is None:
            logger_fn(f"  → Skipping '{sheet.title}': no Description/Rate headers.")
            continue
        for rv in sheet.iter_rows(min_row=header_row_idx+1, max_row=sheet.max_row, max_col=num_cols, values_only=True):
            raw_desc = rv[desc_idx]
            raw_rate = rv[rate_idx]
            if not raw_desc or raw_rate in (None, "", 0):
                continue
            parts = [f"{h}: {v}" for h, v in zip(headers, rv) if v not in (None, "")]  
            ctx = " | ".join(parts)
            texts.append(preprocess_text(ctx))
            descs.append(str(raw_desc).strip())
            rates.append(float(raw_rate))
            cats.append(rv[cat_idx] if include_taxonomy and cat_idx is not None else None)
            subs.append(rv[sub_idx] if include_taxonomy and sub_idx is not None else None)
        logger_fn(f"  → Loaded {len(rates)} total items so far…")
    wb_price.close()
    if not rates:
        raise RuntimeError("No item descriptions with rates found in pricelist file.")
    logger_fn(f"Total loaded pricelist items: {len(rates)}")
    return texts, descs, rates, cats, subs

def load_inquiry_data(inquiry_path, logger_fn):
    logger_fn("Reading inquiry file...")
    wb_inq = load_workbook(inquiry_path, read_only=False, data_only=False)
    items_to_fill, header_rows = [], {}
    for sheet in wb_inq.worksheets:
        logger_fn(f"Scanning inquiry sheet '{sheet.title}' for headers...")
        desc_col = rate_col = qty_col = None
        hdr_row = None
        for r in range(1, 11):
            for c in range(1, sheet.max_column+1):
                v = sheet.cell(row=r, column=c).value
                if isinstance(v, str):
                    low = v.strip().lower()
                    if low == "description": desc_col = c
                    elif low == "rate": rate_col = c
                    elif low in ("qty", "quantity"): qty_col = c
            if desc_col and rate_col:
                hdr_row = r
                header_rows[sheet] = (hdr_row, desc_col, rate_col, qty_col)
                break
        if hdr_row:
            for r in range(hdr_row+1, sheet.max_row+1):
                desc_cell = sheet.cell(row=r, column=desc_col)
                rate_cell = sheet.cell(row=r, column=rate_col)
                qty_cell  = sheet.cell(row=r, column=qty_col) if qty_col else None
                if not desc_cell.value or str(desc_cell.value).strip() == "": continue
                if qty_col and (qty_cell.value is None or str(qty_cell.value).strip() == ""): continue
                if rate_cell.value not in (None, ""): continue
                items_to_fill.append((rate_cell, preprocess_text(str(desc_cell.value))))
    if not items_to_fill:
        raise RuntimeError("No inquiry items with empty rates found in the inquiry file.")
    return wb_inq, items_to_fill, header_rows

def get_embeddings(client, texts, model, logger_fn, input_type="search_document"):
    embs = []
    for i in range(0, len(texts), EMBEDDING_BATCH_SIZE):
        batch = texts[i:i + EMBEDDING_BATCH_SIZE]
        logger_fn(f"Requesting batch {i//EMBEDDING_BATCH_SIZE + 1} from Cohere...")
        resp = client.embed(
            texts=batch,
            model=model,
            input_type=input_type,
            output_dimension=EMBEDDING_DIMENSION,
            embedding_types=["float"]
        )
        embs.extend(resp.embeddings.float)
        logger_fn("Received embeddings from Cohere.")
    return np.array(embs)

def fill_inquiry_rates(client, wb_inq, items_to_fill,
                       pricelist_texts, pricelist_simple_descs, pricelist_rates,
                       pricelist_cats, pricelist_subs, header_rows,
                       model, logger_fn, use_fuzzy, use_taxonomy):
    for sheet, (hdr, _, _, _) in header_rows.items():
        base = sheet.max_column
        sheet.cell(row=hdr, column=base+1, value="Matched Description")
        sheet.cell(row=hdr, column=base+2, value="Similarity Score")
        if use_taxonomy:
            sheet.cell(row=hdr, column=base+3, value="Category")
            sheet.cell(row=hdr, column=base+4, value="SubCategory")
    logger_fn("Computing embeddings for pricelist items...")
    pl_embeds = get_embeddings(client, pricelist_texts, model, logger_fn, input_type="search_document")
    logger_fn("Computing embeddings for inquiry descriptions...")
    in_texts = [desc for _, desc in items_to_fill]
    in_embeds = get_embeddings(client, in_texts, model, logger_fn, input_type="search_query")
    pl_unit = pl_embeds / np.linalg.norm(pl_embeds, axis=1, keepdims=True)
    in_unit = in_embeds / np.linalg.norm(in_embeds, axis=1, keepdims=True)
    sim_mat = in_unit.dot(pl_unit.T)
    for idx, (rate_cell, _) in enumerate(items_to_fill):
        sims = sim_mat[idx]
        best_idx = int(sims.argmax())
        best_score = float(sims[best_idx])
        if use_fuzzy and best_score < FUZZY_THRESHOLD:
            top_idxs = np.argsort(sims)[-FALLBACK_CANDIDATES:][::-1]
            combined_scores = []
            for j in top_idxs:
                fuzzy_score = fuzz.token_sort_ratio(in_texts[idx], pricelist_simple_descs[j]) / 100.0
                combined = 0.7 * sims[j] + 0.3 * fuzzy_score
                combined_scores.append((combined, j))
            best_score, best_idx = max(combined_scores, key=lambda x: x[0])
            best_score = float(best_score)
        best_desc = pricelist_simple_descs[best_idx]
        best_rate = float(pricelist_rates[best_idx])
        sheet = rate_cell.parent
        matched_c = sheet.max_column - (2 + (2 if use_taxonomy else 0)) + 1
        score_c = matched_c + 1
        rate_cell.value = best_rate
        rate_cell.number_format = 'General'
        sheet.cell(row=rate_cell.row, column=matched_c).value = best_desc
        sheet.cell(row=rate_cell.row, column=score_c).value = round(best_score, 3)
        if use_taxonomy:
            sheet.cell(row=rate_cell.row, column=score_c+1).value = pricelist_cats[best_idx]
            sheet.cell(row=rate_cell.row, column=score_c+2).value = pricelist_subs[best_idx]
    logger_fn("All items processed. Matches and rates filled in.")

if __name__ == "__main__":
    root = tk.Tk()
    app = PricelistMatcherApp(root)
    root.mainloop()
